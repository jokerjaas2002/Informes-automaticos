import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import matplotlib.pyplot as plt

# Cargar datos desde un archivo CSV
try:
    data = pd.read_csv('datos.csv')
except FileNotFoundError:
    print("Error: El archivo 'datos.csv' no se encontró.")
    exit()

# Realizar un análisis más completo
informe = data.groupby('Categoría')['Ventas'].agg(['sum', 'mean', 'count']).reset_index()
informe.columns = ['Categoría', 'Total Ventas', 'Promedio Ventas', 'Número de Transacciones']

# Crear un nuevo libro de trabajo
wb = Workbook()
ws = wb.active

# Agregar encabezados
ws.append(['Categoría', 'Total Ventas', 'Promedio Ventas', 'Número de Transacciones'])
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

# Agregar datos
for index, row in informe.iterrows():
    ws.append([row['Categoría'], row['Total Ventas'], row['Promedio Ventas'], row['Número de Transacciones']])

# Ajustar el ancho de las columnas
for column in ws.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column[0].column_letter].width = adjusted_width

# Guardar el archivo
wb.save('informe_ventas.xlsx')
print("Informe guardado como 'informe_ventas.xlsx'")

# Crear un gráfico de barras
plt.figure(figsize=(10, 6))
plt.bar(informe['Categoría'], informe['Total Ventas'], color='skyblue')
plt.title('Total Ventas por Categoría')
plt.xlabel('Categoría')
plt.ylabel('Total Ventas')
plt.xticks(rotation=45)

# Agregar etiquetas de datos en las barras
for index, value in enumerate(informe['Total Ventas']):
    plt.text(index, value, str(value), ha='center', va='bottom')

plt.tight_layout()

# Guardar el gráfico como imagen
plt.savefig('grafico_ventas.png')
plt.savefig('grafico_ventas.pdf')  # Guardar también en formato PDF
plt.show()